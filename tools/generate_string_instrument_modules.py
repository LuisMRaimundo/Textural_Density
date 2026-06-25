#!/usr/bin/env python3
"""Generate violino/viola/violoncelo/contrabaixo GPR modules from AcousticTable workbooks.

Deprecated: use ``tools/generate_instrument_modules.py`` (English module names + flute).
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402
from microtonal import note_to_midi  # noqa: E402
from utils.notes import normalize_note_string  # noqa: E402

CONFIGS = [
    {
        "module": "violino",
        "display": "Violin",
        "workbook": Path(r"D:\CORDAS\VIOLIN_Zenodo_collections_media.xlsx"),
        "doc_anchor": "violin-violino",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "viola",
        "display": "Viola",
        "workbook": Path(r"D:\CORDAS\ViOLA_Zenodo_collections_media.xlsx"),
        "doc_anchor": "viola",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "violoncelo",
        "display": "Cello",
        "workbook": Path(r"D:\CORDAS\CELLO_Zenodo_collections_media.xlsx"),
        "doc_anchor": "cello-violoncelo",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "contrabaixo",
        "display": "Double bass",
        "workbook": Path(r"D:\CORDAS\DOUBLEBASS_Zenodo_collections_media.xlsx"),
        "doc_anchor": "double-bass-contrabaixo",
        "technique": "arco / ordinario sustains",
    },
]

GPR_BODY = '''
def predict_intermediate_dynamics(pitches, pp_values, mf_values, ff_values):
    """Predict intermediate dynamics using Gaussian Process Regression."""
    from instrumentos.gpr_dynamic_interpolation import predict_intermediate_dynamics_gpr

    return predict_intermediate_dynamics_gpr(pp_values, mf_values, ff_values, logger=logger)
'''


def load_workbook_metadata(workbook: Path, instrument_id: str) -> dict[str, object]:
    """Read Provenance + Registry rows written by populate_td_importer_sheets_from_zenodo_media."""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    meta: dict[str, object] = {}
    try:
        prov = wb["Provenance"]
        prov_header = [c for c in next(prov.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in prov.iter_rows(min_row=2, values_only=True):
            if row and row[0] == instrument_id:
                meta.update(dict(zip(prov_header, row)))
                break
        reg = wb["Registry"]
        reg_header = [c for c in next(reg.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in reg.iter_rows(min_row=2, values_only=True):
            if row and row[0] == instrument_id:
                reg_map = dict(zip(reg_header, row))
                meta["sounding_range_low_midi"] = int(reg_map["sounding_range_low_midi"])
                meta["sounding_range_high_midi"] = int(reg_map["sounding_range_high_midi"])
                break
    finally:
        wb.close()
    return meta


def load_spectral_data(workbook: Path) -> dict[str, dict[str, float]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["AcousticTable"]
    raw: dict[str, dict[str, float]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        note = normalize_note_string(str(row[1]))
        raw[note][str(row[3])] = round(float(row[4]), 6)
    wb.close()
    return dict(raw)


def render_module(cfg: dict, spectral: dict[str, dict[str, float]], meta: dict[str, object]) -> str:
    notes = sorted(spectral.keys(), key=lambda n: note_to_midi(n))
    table_lo = int(note_to_midi(notes[0]))
    table_hi = int(note_to_midi(notes[-1]))
    lo = int(meta.get("sounding_range_low_midi", table_lo))
    hi = int(meta.get("sounding_range_high_midi", table_hi))
    citation = str(
        meta.get("citation")
        or (
            f"Sparse {cfg['display'].lower()} CDM table from IOWA and ORCH arco sustain collections; "
            "midpoint summary at pp/mf/ff (Zenodo curation workbook)."
        )
    )
    source_id = str(
        meta.get("source_url_or_identifier")
        or f"docs/instrument_acoustic_sources.md#{cfg['doc_anchor']}"
    )
    transform_notes = str(meta.get("transform_parameters") or "").strip()
    extraction = (
        "Combined Density Metric midpoint of IOWA/ORCH collections "
        f"({transform_notes}); GPR interpolation by pitch/dynamic"
        if transform_notes
        else (
            "Combined Density Metric midpoint of IOWA/ORCH collections; "
            "GPR interpolation by pitch/dynamic"
        )
    )
    version = str(meta.get("import_date") or "2026-06-19")
    table_lines = [
        f"    '{note}': {{'pp': {spectral[note]['pp']}, 'mf': {spectral[note]['mf']}, 'ff': {spectral[note]['ff']}}},"
        for note in notes
    ]
    table = "\n".join(table_lines)
    mod = cfg["module"]
    display = cfg["display"]
    return f'''# instrumentos/{mod}.py
"""
{display} instrument density module.

The ``spectral_data`` table stores sparse Combined Density Metric (CDM) values
from **external acoustic sources** (IOWA + ORCH arco sustain collections,
midpoint summary at pp/mf/ff). Intermediate dynamics are interpolated via GPR.

Runtime analysis does not ingest audio; it maps notated pitch + dynamic to these
pre-loaded acoustic metadata tables.
"""

from instrumentos.provenance import InstrumentSource

INSTRUMENT_SOURCE = InstrumentSource(
    source_type="external_acoustic_metadata",
    citation=(
        "{citation}"
    ),
    source_url_or_identifier={source_id!r},
    extraction_method=(
        "{extraction}"
    ),
    dynamic_levels=("pp", "mf", "ff"),
    pitch_range=({lo}, {hi}),
    uncertainty="medium",
    version="{version}",
)

import logging

from utils.notes import normalize_note_string

logger = logging.getLogger("{mod}")

# CDM medians: (IOWA + ORCH) / 2 per note at pp, mf, ff ({cfg["technique"]}).
spectral_data = {{
{table}
}}


def calcular_densidade(nota, dinamica):
    """Compute density from spectral CDM table (MIDI-space lookup, octave-safe)."""
    from instrumentos.spectral_lookup import lookup_spectral_density

    return lookup_spectral_density(
        spectral_data,
        nota,
        dinamica,
        logger=logger,
        preprocess=normalize_note_string,
    )

{GPR_BODY.strip()}
'''


def main() -> int:
    for cfg in CONFIGS:
        spectral = load_spectral_data(cfg["workbook"])
        meta = load_workbook_metadata(cfg["workbook"], cfg["module"])
        out = ROOT / "instrumentos" / f"{cfg['module']}.py"
        out.write_text(render_module(cfg, spectral, meta), encoding="utf-8")
        notes = sorted(spectral.keys(), key=lambda n: note_to_midi(n))
        print(
            f"Wrote {out.name}: {len(notes)} notes, "
            f"MIDI {int(note_to_midi(notes[0]))}-{int(note_to_midi(notes[-1]))}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
