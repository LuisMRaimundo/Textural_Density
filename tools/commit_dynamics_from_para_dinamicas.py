#!/usr/bin/env python3
"""Commit 10-dynamic CDM ladders from Desktop ``para dinâmicas`` Dynamics10 books.

Reads each ``*_Dynamics10.xlsx`` ``Results`` sheet (status=ok rows only) and
writes ``instrumentos/<module>.py``. Note labels are collapsed with
``normalize_media_note_label`` (``Bb1`` → ``A#1``, ``F4 (2)`` → ``F4``).
Interior cells are clamped into their measured [pp, mf] / [mf, ff] segment
to satisfy the pitched-ladder hygiene contract.

    python tools/commit_dynamics_from_para_dinamicas.py
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from microtonal import note_to_midi_strict  # noqa: E402
from utils.notes import normalize_media_note_label  # noqa: E402

from tools.generate_full_dynamics_modules_from_xlsx import (  # noqa: E402
    InstrumentSpec,
    render_module as render_wind_module,
)
from tools.generate_violin_technique_modules_from_ok_workbooks import (  # noqa: E402
    DYNAMIC_LEVELS,
    _clamp_interiors,
    render_module as render_string_module,
)

VERSION = "2026-09-03"
DEFAULT_SHEET = "Results"


def _find_source_folder() -> Path:
    desk = Path(r"C:\Users\lmr20\Desktop")
    matches = [
        p
        for p in desk.iterdir()
        if p.is_dir() and p.name.lower().startswith("para din")
    ]
    if not matches:
        raise SystemExit(f"Desktop folder starting with 'para din' not found under {desk}")
    return matches[0]


@dataclass(frozen=True)
class Job:
    xlsx_name: str
    family: str  # "wind" | "string"
    spec: dict


JOBS: tuple[Job, ...] = (
    # Woodwinds (ordinary sustain)
    Job(
        "Flute_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "flute",
            "display": "Flute",
            "doc_anchor": "flute",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Clarinet_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "clarinet",
            "display": "Clarinet",
            "doc_anchor": "clarinet",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Oboe_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "oboe",
            "display": "Oboe",
            "doc_anchor": "oboe",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Basson_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "bassoon",
            "display": "Bassoon",
            "doc_anchor": "bassoon",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Piccolo_Zenodo_collections_ordinario_Dynamics10.xlsx",
        "wind",
        {
            "module": "piccolo",
            "display": "Piccolo",
            "doc_anchor": "piccolo",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "English_horn_Zenodo_collections_ordinario_Dynamics10.xlsx",
        "wind",
        {
            "module": "english_horn",
            "display": "English horn",
            "doc_anchor": "english-horn",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Bass_clarinet_in_Bb_Zenodo_collections_ordinario_Dynamics10.xlsx",
        "wind",
        {
            "module": "bass_clarinet",
            "display": "Bass clarinet",
            "doc_anchor": "bass-clarinet",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Contrabassoon_Zenodo_collections_ordinario_Dynamics10.xlsx",
        "wind",
        {
            "module": "contrabassoon",
            "display": "Contrabassoon",
            "doc_anchor": "contrabassoon",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    # Brass
    Job(
        "Horn_in_F_Zenodo_collections_ordinario_Dynamics10.xlsx",
        "wind",
        {
            "module": "horn",
            "display": "Horn",
            "doc_anchor": "horn",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Trumpet_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "trumpet",
            "display": "Trumpet",
            "doc_anchor": "trumpet",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Trombone_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "trombone",
            "display": "Trombone",
            "doc_anchor": "trombone",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    Job(
        "Tuba_Zenodo_collections_Dynamics10.xlsx",
        "wind",
        {
            "module": "tuba",
            "display": "Tuba",
            "doc_anchor": "tuba",
            "source_technique": "ordinary_sustain",
            "family_blurb": "ordinary sustain",
        },
    ),
    # Strings — ordinary + techniques (no sul tasto in this folder)
    Job(
        "VIOLIN_Zenodo_collections_Arco_normal_Dynamics10.xlsx",
        "string",
        {
            "module": "violin",
            "instrument_label": "Violin",
            "technique_label": "arco ordinario",
            "source_technique": "arco_sustain",
            "doc_anchor": "violin",
            "uncertainty": "medium",
            "citation_pool": (
                "dest Zenodo Violin_Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Violin_Zenodo_collections_con_sordino_Dynamics10.xlsx",
        "string",
        {
            "module": "violin_sordina",
            "instrument_label": "Violin",
            "technique_label": "arco con sordino",
            "source_technique": "arco_sordina",
            "doc_anchor": "violin-sordina",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Violin_con sordino Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Violin_Zenodo_collections_sul_ponticello_Dynamics10.xlsx",
        "string",
        {
            "module": "violin_sul_ponticello",
            "instrument_label": "Violin",
            "technique_label": "arco sul ponticello",
            "source_technique": "arco_sul_ponticello",
            "doc_anchor": "violin-sul-ponticello",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Violin_sul ponticello Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Violin_Zenodo_collections_harmonics_Dynamics10.xlsx",
        "string",
        {
            "module": "violin_harmonics",
            "instrument_label": "Violin",
            "technique_label": "arco harmonics",
            "source_technique": "arco_harmonic",
            "doc_anchor": "violin-harmonics",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Violin_harmonics Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "VIOLA_Zenodo_collections_Arco_normal_Dynamics10.xlsx",
        "string",
        {
            "module": "viola",
            "instrument_label": "Viola",
            "technique_label": "arco ordinario",
            "source_technique": "arco_sustain",
            "doc_anchor": "viola",
            "uncertainty": "medium",
            "citation_pool": (
                "dest Zenodo VIOLA_Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Viola_Zenodo_collections_con_sordino_Dynamics10.xlsx",
        "string",
        {
            "module": "viola_sordina",
            "instrument_label": "Viola",
            "technique_label": "arco con sordino",
            "source_technique": "arco_sordina",
            "doc_anchor": "viola-sordina",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Viola_con sordino Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Viola_Zenodo_collections_sul_ponticello_Dynamics10.xlsx",
        "string",
        {
            "module": "viola_sul_ponticello",
            "instrument_label": "Viola",
            "technique_label": "arco sul ponticello",
            "source_technique": "arco_sul_ponticello",
            "doc_anchor": "viola-sul-ponticello",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Viola_sul ponticello Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Viola_Zenodo_collections_harmonics_Dynamics10.xlsx",
        "string",
        {
            "module": "viola_harmonics",
            "instrument_label": "Viola",
            "technique_label": "arco harmonics",
            "source_technique": "arco_harmonic",
            "doc_anchor": "viola-harmonics",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Viola_harmonics Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "CELLO_Zenodo_collections_Arco_normal_Dynamics10.xlsx",
        "string",
        {
            "module": "cello",
            "instrument_label": "Cello",
            "technique_label": "arco ordinario",
            "source_technique": "arco_sustain",
            "doc_anchor": "cello",
            "uncertainty": "medium",
            "citation_pool": (
                "dest Zenodo Cello_Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Cello_Zenodo_collections_con_sordino_Dynamics10.xlsx",
        "string",
        {
            "module": "cello_sordina",
            "instrument_label": "Cello",
            "technique_label": "arco con sordino",
            "source_technique": "arco_sordina",
            "doc_anchor": "cello-sordina",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Cello_con sordino Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Cello_Zenodo_collections_sul_ponticello_Dynamics10.xlsx",
        "string",
        {
            "module": "cello_sul_ponticello",
            "instrument_label": "Cello",
            "technique_label": "arco sul ponticello",
            "source_technique": "arco_sul_ponticello",
            "doc_anchor": "cello-sul-ponticello",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Cello_sul ponticello Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Cello_Zenodo_collections_harmonics_Dynamics10.xlsx",
        "string",
        {
            "module": "cello_harmonics",
            "instrument_label": "Cello",
            "technique_label": "arco harmonics",
            "source_technique": "arco_harmonic",
            "doc_anchor": "cello-harmonics",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo Cello_harmonics Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "DOUBLEBASS_Zenodo_collections_Arco_normal_Dynamics10.xlsx",
        "string",
        {
            "module": "double_bass",
            "instrument_label": "Double bass",
            "technique_label": "arco ordinario",
            "source_technique": "arco_sustain",
            "doc_anchor": "double-bass-double_bass",
            "uncertainty": "medium",
            "citation_pool": (
                "dest Zenodo DBass_Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Double_bass_Zenodo_collections_con_sordino_Dynamics10.xlsx",
        "string",
        {
            "module": "double_bass_sordina",
            "instrument_label": "Double bass",
            "technique_label": "arco con sordino",
            "source_technique": "arco_sordina",
            "doc_anchor": "double-bass-sordina",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo DoubleBass_con sordino Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Double_bass_Zenodo_collections_sul_ponticello_Dynamics10.xlsx",
        "string",
        {
            "module": "double_bass_sul_ponticello",
            "instrument_label": "Double bass",
            "technique_label": "arco sul ponticello",
            "source_technique": "arco_sul_ponticello",
            "doc_anchor": "double-bass-sul-ponticello",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo DoubleBass_sul ponticello Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
    Job(
        "Double_bass_Zenodo_collections_harmonics_Dynamics10.xlsx",
        "string",
        {
            "module": "double_bass_harmonics",
            "instrument_label": "Double bass",
            "technique_label": "arco harmonics",
            "source_technique": "arco_harmonic",
            "doc_anchor": "double-bass-harmonics",
            "uncertainty": "high",
            "citation_pool": (
                "dest Zenodo DoubleBass_harmonics Media (IOWA+Orchidea average); "
                "Dynamics_predicter Results ladder"
            ),
        },
    ),
)


def load_results_ladder(path: Path, sheet: str = DEFAULT_SHEET) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Sheet {sheet!r} not in {path}; have {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit(f"Empty sheet {sheet!r} in {path}")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    for required in ("note", *DYNAMIC_LEVELS):
        if required not in col:
            raise SystemExit(f"Missing column {required!r} on {path.name} / {sheet}")

    ladder: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if not row or row[col["note"]] is None:
            continue
        status = row[col["status"]] if "status" in col else "ok"
        if status is not None and str(status).strip().lower() not in ("ok", ""):
            continue
        values = {dyn: row[col[dyn]] for dyn in DYNAMIC_LEVELS}
        if any(not isinstance(v, (int, float)) for v in values.values()):
            continue
        note = normalize_media_note_label(str(row[col["note"]]).strip())
        dyn_row = {dyn: float(v) for dyn, v in values.items()}
        _clamp_interiors(note, dyn_row)
        if note in ladder:
            raise SystemExit(f"Duplicate note {note} in {path.name}")
        ladder[note] = dyn_row

    if not ladder:
        raise SystemExit(f"No usable ok rows in {path}")
    return dict(sorted(ladder.items(), key=lambda kv: note_to_midi_strict(kv[0])))


def _write_wind(job: Job, table: dict[str, dict[str, float]], source: Path) -> str:
    spec = InstrumentSpec(
        module=job.spec["module"],
        display=job.spec["display"],
        xlsx_name=source.name,
        doc_anchor=job.spec["doc_anchor"],
        source_technique=job.spec["source_technique"],
        family_blurb=job.spec["family_blurb"],
    )
    text = render_wind_module(spec, table, source_path=source, sheet=DEFAULT_SHEET)
    text = text.replace('version="2026-08-03"', f'version="{VERSION}"')
    return text


def _write_string(job: Job, table: dict[str, dict[str, float]], source: Path) -> str:
    rounded = {note: {d: round(v, 6) for d, v in row.items()} for note, row in table.items()}
    spec = {
        **job.spec,
        "workbook": source.name,
        "version": VERSION,
    }
    return render_string_module(spec, rounded)


def main() -> int:
    folder = _find_source_folder()
    summary: list[dict] = []
    print(f"Folder: {folder}")
    print(f"Jobs: {len(JOBS)}  version={VERSION}")
    for job in JOBS:
        src = folder / job.xlsx_name
        if not src.is_file():
            raise SystemExit(f"Missing workbook: {src}")
        table = load_results_ladder(src)
        if job.family == "wind":
            text = _write_wind(job, table, src)
        else:
            text = _write_string(job, table, src)
        out = ROOT / "instrumentos" / f"{job.spec['module']}.py"
        out.write_text(text, encoding="utf-8")
        notes = list(table)
        midis = [int(note_to_midi_strict(n)) for n in notes]
        probes = {}
        for probe in ("C2", "E1", "F1", "C3", "C4", "G4", "A3", "A#4", "C5", "G5", "B7", "C6"):
            if probe in table:
                probes[probe] = {d: table[probe][d] for d in ("pp", "mf", "ff")}
        row = {
            "module": job.spec["module"],
            "xlsx": job.xlsx_name,
            "n": len(table),
            "first": notes[0],
            "last": notes[-1],
            "midi": [min(midis), max(midis)],
            "probes": probes,
        }
        summary.append(row)
        print(
            f"  {out.name:32} n={len(table):3}  {notes[0]}..{notes[-1]}  "
            f"MIDI {min(midis)}-{max(midis)}  from {src.name}"
        )
    log_path = folder / "Dynamics10_td_commit_summary.json"
    log_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    (ROOT / "tools" / "_dynamics10_commit_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print(f"Wrote {len(summary)} modules. Summary: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
