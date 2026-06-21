#!/usr/bin/env python3
"""Generate GPR instrument modules from Zenodo AcousticTable workbooks."""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402
from microtonal import note_to_midi  # noqa: E402
from utils.notes import normalize_note_string  # noqa: E402

CONFIGS = [
    {
        "module": "flute",
        "display": "Flute",
        "workbook": Path(r"D:\MADEIRAS\Flute_Zenodo_collections_media.xlsx"),
        "registry_id": "flute",
        "technique": "sustains (ordinario)",
    },
    {
        "module": "violin",
        "display": "Violin",
        "workbook": Path(r"D:\CORDAS\VIOLIN_Zenodo_collections_media.xlsx"),
        "registry_id": "violin",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "viola",
        "display": "Viola",
        "workbook": Path(r"D:\CORDAS\ViOLA_Zenodo_collections_media.xlsx"),
        "registry_id": "viola",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "cello",
        "display": "Cello",
        "workbook": Path(r"D:\CORDAS\CELLO_Zenodo_collections_media.xlsx"),
        "registry_id": "cello",
        "technique": "arco / ordinario sustains",
    },
    {
        "module": "double_bass",
        "display": "Double bass",
        "workbook": Path(r"D:\CORDAS\DOUBLEBASS_Zenodo_collections_media.xlsx"),
        "registry_id": "double_bass",
        "technique": "arco / ordinario sustains",
    },
]

GPR_BODY = '''
def predict_intermediate_dynamics(pitches, pp_values, mf_values, ff_values):
    """Predict intermediate dynamics using Gaussian Process Regression."""
    dynamic_levels = {
        "pppp": 1, "ppp": 2, "pp": 3, "p": 4, "mf": 5,
        "f": 6, "ff": 7, "fff": 8, "ffff": 9,
    }
    all_dynamics = list(dynamic_levels.keys())
    predictions = {dynamic: [] for dynamic in all_dynamics}

    existing_levels = np.array([dynamic_levels[d] for d in ["pp", "mf", "ff"]]).reshape(-1, 1)
    all_levels = np.array([dynamic_levels[d] for d in all_dynamics]).reshape(-1, 1)

    try:
        y_train = np.array([pp_values, mf_values, ff_values]).T
        if y_train.size == 0 or np.isnan(y_train).any():
            logger.warning("Insufficient or invalid training data for GPR")
            return {d: np.zeros_like(pp_values) for d in all_dynamics}

        matern_kernel = C(1.0) * Matern(length_scale=1.0, nu=1.5)
        gpr = GaussianProcessRegressor(kernel=matern_kernel, n_restarts_optimizer=10, alpha=1e-1)

        for i, y in enumerate(y_train):
            gpr.fit(existing_levels, y)
            y_pred = gpr.predict(all_levels)
            for j, dynamic in enumerate(all_dynamics):
                predictions[dynamic].append(y_pred[j])

        return {k: np.array(v) for k, v in predictions.items()}
    except Exception as e:
        logger.error(f"Error predicting intermediate dynamics: {e}")
        return {d: np.zeros_like(pp_values) for d in all_dynamics}
'''


def load_workbook_metadata(workbook: Path, instrument_id: str) -> dict[str, object]:
    """Read Provenance + Registry rows written by populate_td_importer_sheets_from_zenodo_media."""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    meta: dict[str, object] = {}
    try:
        if "Provenance" in wb.sheetnames:
            prov = wb["Provenance"]
            prov_header = [c for c in next(prov.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row in prov.iter_rows(min_row=2, values_only=True):
                if row and row[0] == instrument_id:
                    meta.update(dict(zip(prov_header, row)))
                    break
        if "Registry" in wb.sheetnames:
            reg = wb["Registry"]
            reg_header = [c for c in next(reg.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row in reg.iter_rows(min_row=2, values_only=True):
                if row and row[0] == instrument_id:
                    reg_map = dict(zip(reg_header, row))
                    meta["sounding_range_low_midi"] = int(reg_map["sounding_range_low_midi"])
                    meta["sounding_range_high_midi"] = int(reg_map["sounding_range_high_midi"])
                    break
    finally:
        wb.close()
    return meta


def load_spectral_data(workbook: Path) -> dict[str, dict[str, float]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["AcousticTable"]
    raw: dict[str, dict[str, float]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        note = normalize_note_string(str(row[1]))
        raw[note][str(row[3])] = round(float(row[4]), 6)
    wb.close()
    return dict(raw)


def render_module(cfg: dict, spectral: dict[str, dict[str, float]], meta: dict[str, object]) -> str:
    notes = sorted(spectral.keys(), key=lambda n: note_to_midi(n))
    table_lo = int(note_to_midi(notes[0]))
    table_hi = int(note_to_midi(notes[-1]))
    lo = int(meta.get("sounding_range_low_midi", table_lo))
    hi = int(meta.get("sounding_range_high_midi", table_hi))
    citation = str(
        meta.get("citation")
        or (
            f"Sparse {cfg['display'].lower()} CDM table from IOWA and ORCH sustain collections; "
            "midpoint summary at pp/mf/ff (Zenodo curation workbook)."
        )
    )
    source_id = str(
        meta.get("source_url_or_identifier")
        or f"docs/instrument_acoustic_sources.md#{cfg['module']}"
    )
    transform_notes = str(meta.get("transform_parameters") or "").strip()
    extraction = (
        "Combined Density Metric midpoint of IOWA/ORCH collections "
        f"({transform_notes}); GPR interpolation by pitch/dynamic"
        if transform_notes
        else (
            "Combined Density Metric midpoint of IOWA/ORCH collections; "
            "GPR interpolation by pitch/dynamic"
        )
    )
    version = str(meta.get("import_date") or "2026-06-19")
    table_lines = [
        f"    '{note}': {{'pp': {spectral[note]['pp']}, 'mf': {spectral[note]['mf']}, 'ff': {spectral[note]['ff']}}},"
        for note in notes
    ]
    table = "\n".join(table_lines)
    mod = cfg["module"]
    display = cfg["display"]
    return f'''# instrumentos/{mod}.py
"""
{display} instrument density module.

The ``spectral_data`` table stores sparse Combined Density Metric (CDM) values
from **external acoustic sources** (IOWA + ORCH sustain collections,
midpoint summary at pp/mf/ff). Intermediate dynamics are interpolated via GPR.

Runtime analysis does not ingest audio; it maps notated pitch + dynamic to these
pre-loaded acoustic metadata tables.
"""

from instrumentos.provenance import InstrumentSource

INSTRUMENT_SOURCE = InstrumentSource(
    source_type="external_acoustic_metadata",
    citation=(
        "{citation}"
    ),
    source_url_or_identifier={source_id!r},
    extraction_method=(
        "{extraction}"
    ),
    dynamic_levels=("pp", "mf", "ff"),
    pitch_range=({lo}, {hi}),
    uncertainty="medium",
    version="{version}",
)

import logging

import numpy as np
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import ConstantKernel as C, Matern
from utils.notes import normalize_note_string

logger = logging.getLogger("{mod}")

# CDM medians: (IOWA + ORCH) / 2 per note at pp, mf, ff ({cfg["technique"]}).
spectral_data = {{
{table}
}}


def calcular_densidade(nota, dinamica):
    """Compute density from spectral CDM table (MIDI-space lookup, octave-safe)."""
    from instrumentos.spectral_lookup import lookup_spectral_density

    return lookup_spectral_density(
        spectral_data,
        nota,
        dinamica,
        logger=logger,
        preprocess=normalize_note_string,
    )

{GPR_BODY.strip()}
'''


def main() -> int:
    for cfg in CONFIGS:
        if not cfg["workbook"].is_file():
            print(f"SKIP {cfg['module']}: workbook not found ({cfg['workbook']})")
            continue
        spectral = load_spectral_data(cfg["workbook"])
        meta = load_workbook_metadata(cfg["workbook"], cfg["registry_id"])
        out = ROOT / "instrumentos" / f"{cfg['module']}.py"
        out.write_text(render_module(cfg, spectral, meta), encoding="utf-8")
        notes = sorted(spectral.keys(), key=lambda n: note_to_midi(n))
        print(
            f"Wrote {out.name}: {len(notes)} notes, "
            f"MIDI {int(note_to_midi(notes[0]))}-{int(note_to_midi(notes[-1]))}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
