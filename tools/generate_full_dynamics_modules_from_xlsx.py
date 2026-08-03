#!/usr/bin/env python3
"""Commit full 10-dynamic CDM ladders from Dynamics_predicter Results sheets.

Writes ``instrumentos/<module>.py`` for each configured instrument. Sheet
defaults to ``Results`` (data-faithful imputation).
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from utils.notes import normalize_note_string  # noqa: E402

OUT_DIR = Path(
    r"c:\Users\lmr20\Desktop\Violino - extrapol\Dynamics_predicter\outputs"
)
DEFAULT_SHEET = "Results"

DYNAMICS = (
    "pppp",
    "ppp",
    "pp",
    "p",
    "mp",
    "mf",
    "f",
    "ff",
    "fff",
    "ffff",
)
NOTE_ORDER = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]


@dataclass(frozen=True)
class InstrumentSpec:
    module: str
    display: str
    xlsx_name: str
    doc_anchor: str
    source_technique: str
    family_blurb: str


SPECS: tuple[InstrumentSpec, ...] = (
    InstrumentSpec(
        module="viola",
        display="Viola",
        xlsx_name="Viola_Arco normal_iowa_orchidea_dynamics.xlsx",
        doc_anchor="viola",
        source_technique="arco_sustain",
        family_blurb="arco",
    ),
    InstrumentSpec(
        module="cello",
        display="Cello",
        xlsx_name="Cello_Arco normal_iowa_orchidea_dynamics.xlsx",
        doc_anchor="cello",
        source_technique="arco_sustain",
        family_blurb="arco",
    ),
    InstrumentSpec(
        module="double_bass",
        display="Double bass",
        xlsx_name="DBass_Arco normal__iowa_orchidea_dynamics.xlsx",
        doc_anchor="double-bass-double_bass",
        source_technique="arco_sustain",
        family_blurb="arco",
    ),
    InstrumentSpec(
        module="flute",
        display="Flute",
        xlsx_name="Flute_iowa_orchidea_dynamics.xlsx",
        doc_anchor="flute",
        source_technique="ordinary_sustain",
        family_blurb="ordinary sustain",
    ),
    InstrumentSpec(
        module="clarinet",
        display="Clarinet",
        xlsx_name="Clarinet_iowa_orchidea_dynamics.xlsx",
        doc_anchor="clarinet",
        source_technique="ordinary_sustain",
        family_blurb="ordinary sustain",
    ),
    InstrumentSpec(
        module="bassoon",
        display="Bassoon",
        xlsx_name="Basson_iowa_orchidea_dynamics.xlsx",
        doc_anchor="bassoon",
        source_technique="ordinary_sustain",
        family_blurb="ordinary sustain",
    ),
    InstrumentSpec(
        module="oboe",
        display="Oboe",
        xlsx_name="Oboe_iowa_orchidea_dynamics.xlsx",
        doc_anchor="oboe",
        source_technique="ordinary_sustain",
        family_blurb="ordinary sustain",
    ),
)


def _note_sort_key(note: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-G]#?)(-?\d+)", str(note).strip())
    if not m:
        return (99, 0)
    pc, octv = m.group(1), int(m.group(2))
    return (octv, NOTE_ORDER.index(pc) if pc in NOTE_ORDER else 99)


def _fmt(value: float) -> str:
    return f"{float(value):.9g}"


def load_results_table(path: Path, sheet: str) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} not in {path}; have {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit(f"Empty sheet {sheet!r}")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    for required in ("note", *DYNAMICS):
        if required not in col:
            raise SystemExit(f"Missing column {required!r} on sheet {sheet!r}")

    table: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if not row or row[col["note"]] is None:
            continue
        status = row[col["status"]] if "status" in col else "ok"
        if status is not None and str(status).strip().lower() not in ("ok", ""):
            continue
        note = normalize_note_string(str(row[col["note"]]).strip())
        dyn_row: dict[str, float] = {}
        for dyn in DYNAMICS:
            val = row[col[dyn]]
            if val is None:
                raise SystemExit(f"Missing {dyn} for note {note} in {path.name}")
            dyn_row[dyn] = float(val)
        table[note] = dyn_row

    if not table:
        raise SystemExit(f"No usable rows in {path}")
    return dict(sorted(table.items(), key=lambda kv: _note_sort_key(kv[0])))


def render_module(
    spec: InstrumentSpec,
    table: dict[str, dict[str, float]],
    *,
    source_path: Path,
    sheet: str,
) -> str:
    from microtonal import note_to_midi_strict

    midis = [int(round(note_to_midi_strict(note))) for note in table]
    lo, hi = min(midis), max(midis)
    dyn_tuple = ", ".join(repr(d) for d in DYNAMICS)
    tech = spec.source_technique

    lines = [
        f"# instrumentos/{spec.module}.py",
        '"""',
        f"{spec.display} instrument density module.",
        "",
        "The ``spectral_data`` table stores a committed 10-dynamic Combined Density",
        "Metric (CDM) ladder from the Dynamics_predicter ``Results`` sheet",
        f"(IOWA + ORCH {spec.family_blurb} anchors at pp/mf/ff; remaining levels",
        "committed from that workbook — no runtime dynamic extrapolation).",
        "",
        "Runtime analysis does not ingest audio; it maps notated pitch + dynamic to",
        "these pre-loaded acoustic metadata tables.",
        '"""',
        "",
        "from instrumentos.provenance import InstrumentSource",
        "",
        "INSTRUMENT_SOURCE = InstrumentSource(",
        '    source_type="external_acoustic_metadata",',
        "    citation=(",
        f'        "{spec.display} CDM ladder: IOWA+ORCH measured pp/mf/ff anchors with "',
        '        "committed Dynamics_predicter Results sheet values for all 10 "',
        '        "dynamic levels (not re-extrapolated at runtime)."',
        "    ),",
        f"    source_url_or_identifier='docs/instrument_acoustic_sources.md#{spec.doc_anchor}',",
        "    extraction_method=(",
        f'        "Committed full dynamic ladder from {source_path.name} / sheet {sheet!r}; "',
        '        "pitch lookup via MIDI-space spectral_lookup"',
        "    ),",
        f"    dynamic_levels=({dyn_tuple}),",
        f"    pitch_range=({lo}, {hi}),",
        '    uncertainty="medium",',
        '    version="2026-08-03",',
        f'    source_technique="{tech}",',
        f'    table_supported_techniques=("{tech}",),',
        ")",
        "",
        "import logging",
        "",
        "from utils.notes import normalize_note_string",
        "",
        f'logger = logging.getLogger("{spec.module}")',
        "",
        "# Full 10-dynamic CDM ladder (Results sheet). Anchors pp/mf/ff match measured",
        "# IOWA+ORCH midpoints; other levels are workbook-committed.",
        "spectral_data = {",
    ]
    for note, dyns in table.items():
        parts = [f"{d!r}: {_fmt(dyns[d])}" for d in DYNAMICS]
        lines.append(f"    {note!r}: {{{', '.join(parts)}}},")
    lines.extend(
        [
            "}",
            "",
            "",
            "def calcular_densidade(nota, dinamica):",
            '    """Compute density from spectral CDM table (MIDI-space lookup, octave-safe)."""',
            "    from instrumentos.spectral_lookup import lookup_spectral_density",
            "",
            "    return lookup_spectral_density(",
            "        spectral_data,",
            "        nota,",
            "        dinamica,",
            "        logger=logger,",
            "        preprocess=normalize_note_string,",
            "    )",
            "",
        ]
    )
    return "\n".join(lines)


def generate_one(spec: InstrumentSpec, *, sheet: str, outputs_dir: Path) -> Path:
    xlsx = outputs_dir / spec.xlsx_name
    if not xlsx.is_file():
        raise SystemExit(f"Workbook not found: {xlsx}")
    table = load_results_table(xlsx, sheet)
    out = ROOT / "instrumentos" / f"{spec.module}.py"
    out.write_text(
        render_module(spec, table, source_path=xlsx, sheet=sheet), encoding="utf-8"
    )
    print(f"Wrote {out.name} ({len(table)} pitches × {len(DYNAMICS)} from {xlsx.name})")
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--outputs-dir", type=Path, default=OUT_DIR)
    parser.add_argument(
        "--only",
        nargs="*",
        help="Optional module names to generate (default: all configured)",
    )
    args = parser.parse_args()
    wanted = set(args.only) if args.only else None
    for spec in SPECS:
        if wanted is not None and spec.module not in wanted:
            continue
        generate_one(spec, sheet=args.sheet, outputs_dir=args.outputs_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
