#!/usr/bin/env python3
"""
Import and validate instrument profile workbooks (Excel) into canonical JSON packages.

Excel is the human-facing curation format. Runtime must not read .xlsx directly.
All acoustic table pitches are sounding/concert pitch; transposition is never applied
during import (registry transposition is for score parsing only).

Phase 1a: validate + emit instrumentos/data/*.profile.json — no Python module generation.
"""

from __future__ import annotations

import argparse
import ast
import hashlib
import json
import math
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment-specific
    raise SystemExit(
        "openpyxl is required for the instrument profile Excel importer. "
        "Install with: pip install openpyxl"
    ) from exc

from microtonal import note_to_midi

SCHEMA_VERSION = "1.0.0"
REQUIRED_ACOUSTIC_PITCH_BASIS = "sounding_concert"

ALLOWED_DYNAMICS = frozenset(
    {"pppp", "ppp", "pp", "p", "mp", "mf", "f", "ff", "fff", "ffff"}
)
DYNAMIC_ALIASES = {
    "piano": "pp",
    "forte": "ff",
    "mezzoforte": "mf",
    "mezzo-forte": "mf",
    "mezzopiano": "mp",
}

EMPIRICAL_PROFILE_STATUSES = frozenset(
    {"empirical_source", "empirical_profile"}
)

ALLOWED_CELL_STATUSES = frozenset(
    {
        "measured_proxy",
        "interpolated",
        "manual_entry",
        "scaled_proxy",
        "coarse_default",
        "symbolic_default",
        "missing",
    }
)

ALLOWED_TRANSFORM_POLICIES = frozenset(
    {
        "manual_review_required",
        "identity_v1",
        "linear_rescale_v1",
        "log_rescale_v1",
        "rank_order_only_v1",
        "reject",
    }
)

DEFAULT_TRANSFORM_POLICY = "manual_review_required"

REQUIRED_SHEETS = ("Registry", "AcousticTable", "Provenance")
OPTIONAL_SHEETS = ("Aliases", "ValidationRules", "Notes", "WorkbookMeta")

REGISTRY_COLUMNS = (
    "instrument_id",
    "display_name",
    "family",
    "subfamily",
    "transposition",
    "sounding_range_low_midi",
    "sounding_range_high_midi",
    "comfortable_range_low_midi",
    "comfortable_range_high_midi",
    "profile_status",
    "source_type",
    "uncertainty",
    "source_notes",
    "supported_techniques",
    "aliases",
)

ACOUSTIC_COLUMNS = (
    "instrument_id",
    "note_sounding",
    "midi_sounding",
    "dynamic",
    "value",
    "value_kind",
    "unit",
    "cell_status",
    "source_system",
    "source_file",
    "source_column",
    "source_hash",
    "transform_policy",
    "uncertainty",
    "validation_status",
    "notes",
    "note_written_optional",
    "midi_written_optional",
    "transposition_semitones_optional",
)

PROVENANCE_COLUMNS = (
    "instrument_id",
    "source_type",
    "citation",
    "source_url_or_identifier",
    "upstream_system",
    "upstream_version",
    "analysis_profile_hash",
    "import_run_id",
    "import_date",
    "operator",
    "transform_policy",
    "transform_parameters",
    "rows_accepted",
    "rows_rejected",
    "notes",
)


@dataclass
class ValidationIssue:
    level: str  # "error" | "warning"
    code: str
    message: str
    sheet: str = ""
    row: int | None = None
    instrument_id: str = ""


@dataclass
class ImportResult:
    issues: list[ValidationIssue] = field(default_factory=list)
    packages: dict[str, dict[str, Any]] = field(default_factory=dict)
    import_run_id: str = field(default_factory=lambda: str(uuid.uuid4()))
    workbook_sha256: str = ""

    @property
    def errors(self) -> list[ValidationIssue]:
        return [i for i in self.issues if i.level == "error"]

    @property
    def warnings(self) -> list[ValidationIssue]:
        return [i for i in self.issues if i.level == "warning"]

    @property
    def ok(self) -> bool:
        return not self.errors


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def _sheet_rows(wb, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_normalize_header(h) for h in rows[0]]
    records: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            record[header] = value
        records.append(record)
    return headers, records


def _read_workbook_meta(wb) -> dict[str, str]:
    meta: dict[str, str] = {}
    for sheet_name in ("WorkbookMeta", "Notes", "ValidationRules"):
        headers, records = _sheet_rows(wb, sheet_name)
        if not records:
            continue
        if set(headers) >= {"key", "value"}:
            for rec in records:
                key = _normalize_header(rec.get("key"))
                val = _normalize_header(rec.get("value"))
                if key:
                    meta[key] = val
        elif sheet_name == "WorkbookMeta" and "acoustic_pitch_basis" in headers:
            for rec in records:
                basis = _normalize_header(rec.get("acoustic_pitch_basis"))
                if basis:
                    meta["acoustic_pitch_basis"] = basis
    return meta


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parse_float(value: Any) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(result):
        return None
    return result


def _parse_int(value: Any) -> int | None:
    f = _parse_float(value)
    if f is None:
        return None
    return int(f)


def _normalize_dynamic(raw: Any) -> str | None:
    dyn = str(raw or "").strip().lower()
    if not dyn:
        return None
    if dyn in ALLOWED_DYNAMICS:
        return dyn
    return DYNAMIC_ALIASES.get(dyn)


def _split_pipe(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def _issue(
    issues: list[ValidationIssue],
    level: str,
    code: str,
    message: str,
    *,
    sheet: str = "",
    row: int | None = None,
    instrument_id: str = "",
) -> None:
    issues.append(
        ValidationIssue(
            level=level,
            code=code,
            message=message,
            sheet=sheet,
            row=row,
            instrument_id=instrument_id,
        )
    )


def validate_and_build(
    workbook_path: Path,
    *,
    strict: bool = False,
) -> ImportResult:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    result = ImportResult(workbook_sha256=_sha256_file(workbook_path))

    meta = _read_workbook_meta(wb)
    pitch_basis = meta.get("acoustic_pitch_basis", "").strip()
    if not pitch_basis:
        _issue(
            result.issues,
            "error",
            "missing_acoustic_pitch_basis",
            "WorkbookMeta/Notes must declare acoustic_pitch_basis=sounding_concert",
            sheet="WorkbookMeta",
        )
    elif pitch_basis != REQUIRED_ACOUSTIC_PITCH_BASIS:
        _issue(
            result.issues,
            "error",
            "invalid_acoustic_pitch_basis",
            f"acoustic_pitch_basis must be {REQUIRED_ACOUSTIC_PITCH_BASIS!r} (got {pitch_basis!r})",
            sheet="WorkbookMeta",
        )

    reg_headers, registry_rows = _sheet_rows(wb, "Registry")
    ac_headers, acoustic_rows = _sheet_rows(wb, "AcousticTable")
    prov_headers, provenance_rows = _sheet_rows(wb, "Provenance")
    _, alias_rows = _sheet_rows(wb, "Aliases")

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            _issue(
                result.issues,
                "error",
                "missing_sheet",
                f"Required sheet {sheet_name!r} is missing",
                sheet=sheet_name,
            )

    if ac_headers:
        has_sounding = "note_sounding" in ac_headers and "midi_sounding" in ac_headers
        has_generic = "note" in ac_headers or "midi" in ac_headers
        if not has_sounding and has_generic:
            _issue(
                result.issues,
                "error",
                "generic_pitch_columns",
                "AcousticTable must use note_sounding/midi_sounding, not generic note/midi",
                sheet="AcousticTable",
            )
        elif not has_sounding:
            _issue(
                result.issues,
                "error",
                "missing_sounding_columns",
                "AcousticTable requires note_sounding and midi_sounding columns",
                sheet="AcousticTable",
            )

    registry_by_id: dict[str, dict[str, Any]] = {}
    for row_idx, row in enumerate(registry_rows, start=2):
        iid = _normalize_header(row.get("instrument_id"))
        if not iid:
            _issue(
                result.issues,
                "error",
                "missing_instrument_id",
                "Registry row missing instrument_id",
                sheet="Registry",
                row=row_idx,
            )
            continue
        if iid in registry_by_id:
            _issue(
                result.issues,
                "error",
                "duplicate_registry_id",
                f"Duplicate registry instrument_id {iid!r}",
                sheet="Registry",
                row=row_idx,
                instrument_id=iid,
            )
        source_notes = _normalize_header(row.get("source_notes"))
        profile_status = _normalize_header(row.get("profile_status"))
        if profile_status in EMPIRICAL_PROFILE_STATUSES and not source_notes:
            _issue(
                result.issues,
                "error",
                "empirical_without_source_notes",
                f"profile_status={profile_status!r} requires source_notes",
                sheet="Registry",
                row=row_idx,
                instrument_id=iid,
            )
        registry_by_id[iid] = row

    provenance_by_id: dict[str, dict[str, Any]] = {}
    for row_idx, row in enumerate(provenance_rows, start=2):
        iid = _normalize_header(row.get("instrument_id"))
        if not iid:
            continue
        provenance_by_id[iid] = row

    alias_map: dict[str, list[str]] = {}
    for row in alias_rows:
        iid = _normalize_header(row.get("instrument_id"))
        alias = _normalize_header(row.get("alias"))
        if iid and alias:
            alias_map.setdefault(iid, []).append(alias)

    seen_cells: set[tuple[str, float, str]] = set()
    acoustic_by_instrument: dict[str, list[dict[str, Any]]] = {}

    for row_idx, row in enumerate(acoustic_rows, start=2):
        iid = _normalize_header(row.get("instrument_id"))
        if not iid:
            _issue(
                result.issues,
                "error",
                "missing_instrument_id",
                "AcousticTable row missing instrument_id",
                sheet="AcousticTable",
                row=row_idx,
            )
            continue
        if iid not in registry_by_id:
            _issue(
                result.issues,
                "error" if strict else "warning",
                "unknown_instrument_id",
                f"AcousticTable references unknown instrument_id {iid!r}",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )
            if strict:
                continue

        note_sounding = _normalize_header(row.get("note_sounding"))
        midi_raw = row.get("midi_sounding")
        midi_sounding = _parse_float(midi_raw)

        if not note_sounding and midi_sounding is None:
            _issue(
                result.issues,
                "error",
                "missing_pitch",
                "AcousticTable row requires note_sounding and/or midi_sounding",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )
            continue

        if note_sounding and midi_sounding is None:
            try:
                midi_sounding = float(note_to_midi(note_sounding))
            except Exception:
                midi_sounding = None
        if midi_sounding is not None and not note_sounding:
            _issue(
                result.issues,
                "warning",
                "midi_without_note",
                "midi_sounding provided without note_sounding",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )

        if note_sounding:
            try:
                expected_midi = float(note_to_midi(note_sounding))
            except Exception:
                expected_midi = None
                _issue(
                    result.issues,
                    "error",
                    "invalid_note_sounding",
                    f"Cannot parse note_sounding {note_sounding!r}",
                    sheet="AcousticTable",
                    row=row_idx,
                    instrument_id=iid,
                )
            if expected_midi is not None and midi_sounding is not None:
                if abs(expected_midi - midi_sounding) > 0.01:
                    _issue(
                        result.issues,
                        "error",
                        "pitch_mismatch",
                        f"note_sounding {note_sounding!r} (MIDI {expected_midi}) "
                        f"does not match midi_sounding {midi_sounding}",
                        sheet="AcousticTable",
                        row=row_idx,
                        instrument_id=iid,
                    )

        dynamic = _normalize_dynamic(row.get("dynamic"))
        if not dynamic:
            _issue(
                result.issues,
                "error",
                "unknown_dynamic",
                f"Unknown or missing dynamic {row.get('dynamic')!r}",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )
            continue

        cell_status = _normalize_header(row.get("cell_status")) or "missing"
        if cell_status not in ALLOWED_CELL_STATUSES:
            _issue(
                result.issues,
                "error",
                "invalid_cell_status",
                f"Invalid cell_status {cell_status!r}",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )
            continue

        value = _parse_float(row.get("value"))
        if cell_status != "missing":
            if value is None:
                _issue(
                    result.issues,
                    "error",
                    "non_finite_value",
                    f"Non-finite or missing value for cell_status={cell_status!r}",
                    sheet="AcousticTable",
                    row=row_idx,
                    instrument_id=iid,
                )
                continue
        elif value is not None:
            _issue(
                result.issues,
                "warning",
                "missing_with_value",
                "cell_status=missing but value is present; value ignored",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )

        row_transform = _normalize_header(row.get("transform_policy"))
        if row_transform:
            if row_transform not in ALLOWED_TRANSFORM_POLICIES:
                _issue(
                    result.issues,
                    "error",
                    "invalid_transform_policy",
                    f"Invalid transform_policy {row_transform!r}",
                    sheet="AcousticTable",
                    row=row_idx,
                    instrument_id=iid,
                )
                continue
            transform_policy = row_transform
        else:
            transform_policy = DEFAULT_TRANSFORM_POLICY

        if cell_status == "measured_proxy":
            for req_col in ("source_system", "source_file", "source_column"):
                if not _normalize_header(row.get(req_col)):
                    _issue(
                        result.issues,
                        "error",
                        "measured_proxy_missing_provenance",
                        f"measured_proxy requires {req_col}",
                        sheet="AcousticTable",
                        row=row_idx,
                        instrument_id=iid,
                    )

        if midi_sounding is None:
            continue

        # Hard rule: never apply registry transposition to imported sounding pitch.
        registry_row = registry_by_id.get(iid, {})
        _ = _parse_int(registry_row.get("transposition"))  # documented: score-only

        cell_key = (iid, round(midi_sounding, 4), dynamic)
        if cell_key in seen_cells:
            _issue(
                result.issues,
                "error",
                "duplicate_cell",
                f"Duplicate cell instrument_id={iid!r} midi_sounding={midi_sounding} dynamic={dynamic!r}",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )
            continue
        seen_cells.add(cell_key)

        low = _parse_float(registry_row.get("sounding_range_low_midi"))
        high = _parse_float(registry_row.get("sounding_range_high_midi"))
        if low is not None and high is not None and (midi_sounding < low or midi_sounding > high):
            _issue(
                result.issues,
                "warning",
                "outside_sounding_range",
                f"midi_sounding {midi_sounding} outside sounding range [{low}, {high}]",
                sheet="AcousticTable",
                row=row_idx,
                instrument_id=iid,
            )

        written_trace: dict[str, Any] = {}
        if _normalize_header(row.get("note_written_optional")):
            written_trace["note_written_optional"] = _normalize_header(
                row.get("note_written_optional")
            )
        if row.get("midi_written_optional") is not None and str(
            row.get("midi_written_optional")
        ).strip():
            written_trace["midi_written_optional"] = _parse_float(
                row.get("midi_written_optional")
            )
        if row.get("transposition_semitones_optional") is not None and str(
            row.get("transposition_semitones_optional")
        ).strip():
            written_trace["transposition_semitones_optional"] = _parse_int(
                row.get("transposition_semitones_optional")
            )

        cell_record = {
            "note_sounding": note_sounding,
            "midi_sounding": midi_sounding,
            "dynamic": dynamic,
            "value": value,
            "value_kind": _normalize_header(row.get("value_kind")) or "amplitude_proxy",
            "unit": _normalize_header(row.get("unit")) or "td_amplitude_proxy_v1",
            "cell_status": cell_status,
            "transform_policy": transform_policy,
            "cell_provenance": {
                "source_system": _normalize_header(row.get("source_system")),
                "source_file": _normalize_header(row.get("source_file")),
                "source_column": _normalize_header(row.get("source_column")),
                "source_hash": _normalize_header(row.get("source_hash")),
            },
            "written_traceability": written_trace or None,
            "notes": _normalize_header(row.get("notes")),
        }
        acoustic_by_instrument.setdefault(iid, []).append(cell_record)

    if strict:
        for issue in result.warnings:
            if issue.level == "warning":
                _issue(
                    result.issues,
                    "error",
                    issue.code,
                    issue.message,
                    sheet=issue.sheet,
                    row=issue.row,
                    instrument_id=issue.instrument_id,
                )

    if result.errors:
        return result

    import_ts = datetime.now(timezone.utc).isoformat()
    for iid, reg_row in registry_by_id.items():
        prov_row = provenance_by_id.get(iid, {})
        cells = acoustic_by_instrument.get(iid, [])
        package_aliases = _split_pipe(reg_row.get("aliases"))
        package_aliases.extend(alias_map.get(iid, []))
        package_aliases = sorted(set(package_aliases))

        registry_block = {
            "instrument_id": iid,
            "display_name": _normalize_header(reg_row.get("display_name")),
            "family": _normalize_header(reg_row.get("family")),
            "subfamily": _normalize_header(reg_row.get("subfamily")),
            "transposition": _parse_int(reg_row.get("transposition")) or 0,
            "sounding_range": [
                _parse_float(reg_row.get("sounding_range_low_midi")),
                _parse_float(reg_row.get("sounding_range_high_midi")),
            ],
            "comfortable_range": [
                _parse_float(reg_row.get("comfortable_range_low_midi")),
                _parse_float(reg_row.get("comfortable_range_high_midi")),
            ],
            "profile_status": _normalize_header(reg_row.get("profile_status")),
            "source_type": _normalize_header(reg_row.get("source_type")),
            "uncertainty": _normalize_header(reg_row.get("uncertainty")),
            "source_notes": _normalize_header(reg_row.get("source_notes")),
            "supported_techniques": _split_pipe(reg_row.get("supported_techniques")),
            "aliases": package_aliases,
        }

        entries_map: dict[tuple[str, float], dict[str, Any]] = {}
        for cell in cells:
            key = (cell["note_sounding"], cell["midi_sounding"])
            entry = entries_map.setdefault(
                key,
                {
                    "note_sounding": cell["note_sounding"],
                    "midi_sounding": cell["midi_sounding"],
                    "dynamics": {},
                    "cell_status_by_dynamic": {},
                    "written_traceability": cell.get("written_traceability"),
                },
            )
            if cell["cell_status"] != "missing" and cell.get("value") is not None:
                entry["dynamics"][cell["dynamic"]] = cell["value"]
            entry["cell_status_by_dynamic"][cell["dynamic"]] = cell["cell_status"]

        dynamic_levels = sorted(
            {cell["dynamic"] for cell in cells},
            key=lambda d: list(ALLOWED_DYNAMICS).index(d)
            if d in ALLOWED_DYNAMICS
            else 999,
        )

        rows_accepted = len(cells)
        package = {
            "schema_version": SCHEMA_VERSION,
            "acoustic_pitch_basis": REQUIRED_ACOUSTIC_PITCH_BASIS,
            "instrument_id": iid,
            "registry": registry_block,
            "acoustic_table": {
                "dynamic_levels": dynamic_levels,
                "entries": list(entries_map.values()),
                "cells": cells,
            },
            "provenance": {
                "source_type": _normalize_header(prov_row.get("source_type"))
                or registry_block["source_type"],
                "citation": _normalize_header(prov_row.get("citation")),
                "source_url_or_identifier": _normalize_header(
                    prov_row.get("source_url_or_identifier")
                ),
                "upstream_system": _normalize_header(prov_row.get("upstream_system")),
                "upstream_version": _normalize_header(prov_row.get("upstream_version")),
                "analysis_profile_hash": _normalize_header(
                    prov_row.get("analysis_profile_hash")
                ),
                "transform_policy": _normalize_header(prov_row.get("transform_policy"))
                or DEFAULT_TRANSFORM_POLICY,
                "transform_parameters": _normalize_header(
                    prov_row.get("transform_parameters")
                ),
                "operator": _normalize_header(prov_row.get("operator")),
                "notes": _normalize_header(prov_row.get("notes")),
            },
            "import_audit": {
                "import_run_id": result.import_run_id,
                "import_date": import_ts,
                "source_workbook_sha256": result.workbook_sha256,
                "rows_accepted": rows_accepted,
                "rows_rejected": 0,
                "warnings": [w.message for w in result.warnings if w.instrument_id in ("", iid)],
                "dry_run": False,
            },
        }
        result.packages[iid] = package

    return result


def write_packages(packages: dict[str, dict[str, Any]], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for iid, package in packages.items():
        path = output_dir / f"{iid}.profile.json"
        path.write_text(json.dumps(package, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        written.append(path)
    index = {
        "schema_version": SCHEMA_VERSION,
        "acoustic_pitch_basis": REQUIRED_ACOUSTIC_PITCH_BASIS,
        "packages": sorted(packages.keys()),
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    index_path = output_dir / "index.json"
    index_path.write_text(json.dumps(index, indent=2) + "\n", encoding="utf-8")
    written.append(index_path)
    return written


def build_report(result: ImportResult, *, dry_run: bool) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "import_run_id": result.import_run_id,
        "workbook_sha256": result.workbook_sha256,
        "dry_run": dry_run,
        "ok": result.ok,
        "errors": [
            {
                "code": i.code,
                "message": i.message,
                "sheet": i.sheet,
                "row": i.row,
                "instrument_id": i.instrument_id,
            }
            for i in result.errors
        ],
        "warnings": [
            {
                "code": i.code,
                "message": i.message,
                "sheet": i.sheet,
                "row": i.row,
                "instrument_id": i.instrument_id,
            }
            for i in result.warnings
        ],
        "packages": sorted(result.packages.keys()),
    }


def run_import(
    workbook_path: Path,
    output_dir: Path,
    *,
    dry_run: bool = False,
    strict: bool = False,
    report_path: Path | None = None,
) -> ImportResult:
    result = validate_and_build(workbook_path, strict=strict)
    report = build_report(result, dry_run=dry_run)
    if report_path is not None:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")

    if not result.ok:
        return result

    for pkg in result.packages.values():
        pkg["import_audit"]["dry_run"] = dry_run

    if dry_run:
        return result

    write_packages(result.packages, output_dir)
    return result


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import instrument profile Excel workbook to canonical JSON packages."
    )
    parser.add_argument("--workbook", required=True, type=Path, help="Path to .xlsx workbook")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("instrumentos/data"),
        help="Directory for generated .profile.json files",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate only; do not write profile JSON",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Treat warnings such as unknown instrument_id as errors",
    )
    parser.add_argument("--report", type=Path, help="Write validation report JSON to this path")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    if not args.workbook.is_file():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 2
    result = run_import(
        args.workbook,
        args.output_dir,
        dry_run=args.dry_run,
        strict=args.strict,
        report_path=args.report,
    )
    if result.ok:
        mode = "validated (dry-run)" if args.dry_run else "imported"
        print(f"OK: {mode} {len(result.packages)} package(s)")
        if result.warnings:
            print(f"Warnings: {len(result.warnings)}", file=sys.stderr)
        return 0
    print(f"FAILED: {len(result.errors)} error(s)", file=sys.stderr)
    for err in result.errors:
        loc = f"{err.sheet}:{err.row}" if err.row else err.sheet
        print(f"  [{err.code}] {loc} {err.message}", file=sys.stderr)
    return 1


def module_imports_are_safe() -> bool:
    """True if this module does not import audio/SSA/FFT stacks."""
    path = Path(__file__)
    tree = ast.parse(path.read_text(encoding="utf-8"))
    forbidden = (
        "proc_audio",
        "spectral_analysis",
        "Spectral_Analyser",
        "soundSpectrAnalyse",
        "fft",
        "stft",
    )
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                if any(f in alias.name for f in forbidden):
                    return False
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            if any(f in mod for f in forbidden):
                return False
    return True


if __name__ == "__main__":
    raise SystemExit(main())
