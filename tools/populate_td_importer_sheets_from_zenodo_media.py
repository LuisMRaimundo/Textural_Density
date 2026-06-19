#!/usr/bin/env python3
"""
Add Textural Density Phase-1a importer sheets to Zenodo string-media workbooks.

Reads the per-instrument *_Media sheet (IOWA/ORCH midpoint columns) and writes:
WorkbookMeta, Registry, AcousticTable, Provenance, Aliases.

Optionally repairs instrument-specific README / Zenodo_File_metadata copy-paste errors.
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from microtonal import note_to_midi  # noqa: E402

WORKBOOK_META_COLUMNS = ("key", "value")
REGISTRY_COLUMNS = (
    "instrument_id",
    "display_name",
    "family",
    "subfamily",
    "transposition",
    "sounding_range_low_midi",
    "sounding_range_high_midi",
    "comfortable_range_low_midi",
    "comfortable_range_high_midi",
    "profile_status",
    "source_type",
    "uncertainty",
    "source_notes",
    "supported_techniques",
    "aliases",
)
ACOUSTIC_COLUMNS = (
    "instrument_id",
    "note_sounding",
    "midi_sounding",
    "dynamic",
    "value",
    "value_kind",
    "unit",
    "cell_status",
    "source_system",
    "source_file",
    "source_column",
    "source_hash",
    "transform_policy",
    "uncertainty",
    "validation_status",
    "notes",
    "note_written_optional",
    "midi_written_optional",
    "transposition_semitones_optional",
)
PROVENANCE_COLUMNS = (
    "instrument_id",
    "source_type",
    "citation",
    "source_url_or_identifier",
    "upstream_system",
    "upstream_version",
    "analysis_profile_hash",
    "import_run_id",
    "import_date",
    "operator",
    "transform_policy",
    "transform_parameters",
    "rows_accepted",
    "rows_rejected",
    "notes",
)
ALIAS_COLUMNS = ("instrument_id", "alias", "alias_kind")

TD_SHEETS = ("WorkbookMeta", "Registry", "AcousticTable", "Provenance", "Aliases")

INSTRUMENT_CONFIGS: list[dict[str, Any]] = [
    {
        "workbook": Path(r"D:\CORDAS\VIOLIN_Zenodo_collections_media.xlsx"),
        "media_sheet": "Violin_Media",
        "instrument_id": "violino",
        "display_name": "Violin",
        "family": "strings",
        "subfamily": "bowed_strings",
        "sounding": (55, 88),
        "comfortable": (55, 76),
        "registry_aliases": ("violin",),
        "extra_aliases": (),
        "zenodo_key": "violin",
        "zenodo_filename": "violin_arco_sustains_median_summary_v1.xlsx",
        "source_workbook": "VIOLIN.xlsx",
        "readme_title": "Violin median summary for Zenodo",
    },
    {
        "workbook": Path(r"D:\CORDAS\ViOLA_Zenodo_collections_media.xlsx"),
        "media_sheet": "VIOLA_Media",
        "instrument_id": "viola",
        "display_name": "Viola",
        "family": "strings",
        "subfamily": "bowed_strings",
        "sounding": (48, 76),
        "comfortable": (50, 69),
        "registry_aliases": ("viola",),
        "extra_aliases": (),
        "zenodo_key": "viola",
        "zenodo_filename": "viola_arco_sustains_median_summary_v1.xlsx",
        "source_workbook": "Viola.xlsx",
        "readme_title": "Viola median summary for Zenodo",
    },
    {
        "workbook": Path(r"D:\CORDAS\CELLO_Zenodo_collections_media.xlsx"),
        "media_sheet": "Cello_Media",
        "instrument_id": "violoncelo",
        "display_name": "Cello",
        "family": "strings",
        "subfamily": "bowed_strings",
        "sounding": (36, 72),
        "comfortable": (40, 65),
        "registry_aliases": ("cello", "violoncello"),
        "extra_aliases": (),
        "zenodo_key": "cello",
        "zenodo_filename": "cello_arco_sustains_median_summary_v1.xlsx",
        "source_workbook": "Cello.xlsx",
        "readme_title": "Cello median summary for Zenodo",
    },
    {
        "workbook": Path(r"D:\CORDAS\DOUBLEBASS_Zenodo_collections_media.xlsx"),
        "media_sheet": "DBass_Media",
        "instrument_id": "contrabaixo",
        "display_name": "Double bass",
        "family": "strings",
        "subfamily": "bowed_strings",
        "sounding": (28, 60),
        "comfortable": (31, 55),
        "registry_aliases": ("double_bass", "contrabass", "baixo"),
        "extra_aliases": (),
        "zenodo_key": "double_bass",
        "zenodo_filename": "double_bass_arco_sustains_median_summary_v1.xlsx",
        "source_workbook": "DoubleBass.xlsx",
        "readme_title": "Double bass median summary for Zenodo",
    },
]

MEDIA_DYNAMIC_COLUMNS = (
    ("pp", 3, "Media pp = (IOWA pp + ORCH pp) / 2"),
    ("mf", 6, "Media mf = (IOWA mf + ORCH mf) / 2"),
    ("ff", 9, "Media ff = (IOWA ff + ORCH ff) / 2"),
)


def _write_header(ws: Worksheet, headers: tuple[str, ...]) -> None:
    ws.append(list(headers))


def _remove_sheet_if_exists(wb: Workbook, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def _source_notes(cfg: dict[str, Any]) -> str:
    return (
        f"Sparse arco/sustain Combined Density Metric (CDM) table for {cfg['display_name']} "
        f"at pp/mf/ff. Values are the arithmetic midpoint of IOWA and ORCH collection "
        f"summaries on sheet {cfg['media_sheet']!r} (sounding/concert pitch). "
        "External acoustic proxy metadata for Textural Density — not runtime audio analysis."
    )


def _citation(cfg: dict[str, Any]) -> str:
    return (
        f"Median/midpoint summary of {cfg['display_name'].lower()} arco sustained-note "
        "Combined Density Metrics across IOWA and ORCH sound collections (pp, mf, ff)."
    )


def _zenodo_description(cfg: dict[str, Any]) -> str:
    return (
        f"This workbook provides note-level and dynamic-level median summaries for "
        f"{cfg['display_name'].lower()} arco sustained-note combined-density metrics "
        "derived from two sound collections (IOWA and ORCH), at pp, mf and ff dynamic "
        "levels. It is intended as a compact summary-statistics layer accompanying the "
        "full analysis outputs, logs, segmentation metadata and compiled metric files."
    )


def _read_media_rows(workbook_path: Path, media_sheet: str) -> list[tuple[str, dict[str, float]]]:
    """Read computed Media values (columns are often Excel formulas)."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if media_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Missing media sheet {media_sheet!r} in workbook")
    ws = wb[media_sheet]
    rows: list[tuple[str, dict[str, float]]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        note = row[0]
        if note is None or str(note).strip() == "":
            continue
        note_str = str(note).strip()
        dynamics: dict[str, float] = {}
        for dyn, col_idx, _ in MEDIA_DYNAMIC_COLUMNS:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            try:
                dynamics[dyn] = float(val)
            except (TypeError, ValueError):
                continue
        if dynamics:
            rows.append((note_str, dynamics))
    wb.close()
    return rows


def _build_acoustic_rows(cfg: dict[str, Any], media_rows: list[tuple[str, dict[str, float]]]) -> list[tuple[Any, ...]]:
    wb_name = cfg["workbook"].name
    acoustic: list[tuple[Any, ...]] = []
    for note, dynamics in media_rows:
        midi = float(note_to_midi(note))
        for dyn, _, col_label in MEDIA_DYNAMIC_COLUMNS:
            value = dynamics.get(dyn)
            if value is None:
                continue
            acoustic.append(
                (
                    cfg["instrument_id"],
                    note,
                    midi,
                    dyn,
                    value,
                    "combined_density_metric",
                    "cdm_arco_sustain_v1",
                    "measured_proxy",
                    "IOWA+ORCH string corpus",
                    wb_name,
                    col_label,
                    "",
                    "identity_v1",
                    "medium",
                    "accepted",
                    f"Midpoint of IOWA and ORCH CDM on {cfg['media_sheet']}",
                    "",
                    "",
                    "",
                )
            )
    return acoustic


def _update_readme(wb: Workbook, cfg: dict[str, Any]) -> None:
    if "README" not in wb.sheetnames:
        return
    ws = wb["README"]
    instrument = cfg["display_name"].lower()
    replacements = {
        "Violin median summary for Zenodo": cfg["readme_title"],
        "violin/arco/sustains": f"{instrument}/arco/sustains",
        "for violin arco": f"for {instrument} arco",
    }
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value is None:
            continue
        text = str(cell.value)
        new_text = text
        for old, new in replacements.items():
            new_text = new_text.replace(old, new)
        if new_text != text:
            cell.value = new_text
        if cell.value == "Source workbooks":
            label_cell = row[1]
            if label_cell is not None:
                label_cell.value = f"{cfg['source_workbook']} (colocar todos)"


def _update_zenodo_metadata(wb: Workbook, cfg: dict[str, Any], row_count: int) -> None:
    if "Zenodo_File_metadata" not in wb.sheetnames:
        return
    ws = wb["Zenodo_File_metadata"]
    mapping = {str(r[0].value): r[1] for r in ws.iter_rows(min_row=2) if r[0].value}
    updates = {
        "recommended_filename": cfg["zenodo_filename"],
        "title": _citation(cfg),
        "instrument": cfg["display_name"],
        "source_workbook": cfg["source_workbook"],
        "recommended_zenodo_description": _zenodo_description(cfg),
        "quality_control_note": (
            f"{cfg['display_name']} arco sustain CDM medians from IOWA and ORCH collections; "
            "enharmonic spellings preserved from source labels. Review octave pairing on "
            "extreme register rows before treating as final reference data."
        ),
    }
    for row in ws.iter_rows(min_row=2):
        key = row[0].value
        if key is None:
            continue
        key_str = str(key)
        if key_str in updates:
            row[1].value = updates[key_str]
    if "rows_accepted" not in mapping:
        ws.append(("rows_accepted", row_count))
    if "textural_density_importer" not in mapping:
        ws.append(
            (
                "textural_density_importer",
                "WorkbookMeta/Registry/AcousticTable/Provenance/Aliases sheets added for "
                "tools/import_instrument_profiles_from_excel.py (Phase 1a).",
            )
        )


def populate_workbook(cfg: dict[str, Any], *, backup: bool = True, fix_zenodo: bool = True) -> int:
    path = cfg["workbook"]
    if not path.exists():
        raise FileNotFoundError(path)

    if backup:
        backup_path = path.with_suffix(path.suffix + ".backup")
        if not backup_path.exists():
            shutil.copy2(path, backup_path)

    media_rows = _read_media_rows(path, cfg["media_sheet"])
    wb = load_workbook(path)
    acoustic_rows = _build_acoustic_rows(cfg, media_rows)

    for sheet_name in TD_SHEETS:
        _remove_sheet_if_exists(wb, sheet_name)

    meta_ws = wb.create_sheet("WorkbookMeta")
    _write_header(meta_ws, WORKBOOK_META_COLUMNS)
    meta_ws.append(("acoustic_pitch_basis", "sounding_concert"))
    meta_ws.append(("schema_version", "1.0.0"))
    meta_ws.append(("template", "instrument_profiles_template.xlsx"))
    meta_ws.append(("source_media_sheet", cfg["media_sheet"]))
    meta_ws.append(("curated_for", "Textural Density Phase 1a importer"))
    meta_ws.append(("generated_at", datetime.now(timezone.utc).isoformat()))

    reg_ws = wb.create_sheet("Registry")
    _write_header(reg_ws, REGISTRY_COLUMNS)
    sound_low, sound_high = cfg["sounding"]
    comfort_low, comfort_high = cfg["comfortable"]
    reg_ws.append(
        (
            cfg["instrument_id"],
            cfg["display_name"],
            cfg["family"],
            cfg["subfamily"],
            0,
            sound_low,
            sound_high,
            comfort_low,
            comfort_high,
            "empirical_profile",
            "external_acoustic_metadata",
            "medium",
            _source_notes(cfg),
            "arco|pizzicato|tremolo|harmonic|mute",
            "|".join(cfg["registry_aliases"]),
        )
    )

    ac_ws = wb.create_sheet("AcousticTable")
    _write_header(ac_ws, ACOUSTIC_COLUMNS)
    for row in acoustic_rows:
        ac_ws.append(list(row))

    prov_ws = wb.create_sheet("Provenance")
    _write_header(prov_ws, PROVENANCE_COLUMNS)
    prov_ws.append(
        (
            cfg["instrument_id"],
            "external_acoustic_metadata",
            _citation(cfg),
            str(path),
            "IOWA+ORCH string analysis corpus",
            "v1",
            "",
            "",
            datetime.now(timezone.utc).date().isoformat(),
            "populate_td_importer_sheets_from_zenodo_media.py",
            "identity_v1",
            "CDM midpoint pass-through; no rescaling",
            len(acoustic_rows),
            0,
            f"Derived from {cfg['media_sheet']} Media pp/mf/ff columns",
        )
    )

    alias_ws = wb.create_sheet("Aliases")
    _write_header(alias_ws, ALIAS_COLUMNS)
    for alias in cfg["registry_aliases"]:
        alias_ws.append((cfg["instrument_id"], alias, "registry"))
    for alias in cfg["extra_aliases"]:
        alias_ws.append((cfg["instrument_id"], alias, "common"))

    if fix_zenodo:
        _update_readme(wb, cfg)
        _update_zenodo_metadata(wb, cfg, len(acoustic_rows))

    wb.save(path)
    return len(acoustic_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        action="append",
        help="Specific workbook path (default: all four string configs)",
    )
    parser.add_argument("--no-backup", action="store_true")
    parser.add_argument("--no-fix-zenodo", action="store_true")
    args = parser.parse_args()

    configs = INSTRUMENT_CONFIGS
    if args.workbook:
        selected = {p.resolve() for p in args.workbook}
        configs = [c for c in INSTRUMENT_CONFIGS if c["workbook"].resolve() in selected]
        if not configs:
            raise SystemExit("No matching instrument configs for --workbook paths")

    for cfg in configs:
        count = populate_workbook(
            cfg,
            backup=not args.no_backup,
            fix_zenodo=not args.no_fix_zenodo,
        )
        print(
            f"Updated {cfg['workbook']} — {cfg['instrument_id']}: "
            f"{count} AcousticTable rows ({count // 3} notes × 3 dynamics)"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
