#!/usr/bin/env python3
"""Commit violin arco full 10-dynamic CDM ladder from Dynamics_predicter Results.

Default source:
  Desktop/.../Dynamics_predicter/outputs/iowa_orchidea_dynamics.xlsx
  sheet ``Results`` (data-faithful imputation; pp/mf/ff = measured anchors).

Writes ``instrumentos/violin.py`` with all DYNAMIC_LEVELS present per pitch so
runtime orchestration can skip GPR / adaptive-tail extrapolation for violin.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from utils.notes import normalize_note_string  # noqa: E402

DEFAULT_XLSX = Path(
    r"c:\Users\lmr20\Desktop\Violino - extrapol\Dynamics_predicter\outputs"
    r"\iowa_orchidea_dynamics.xlsx"
)
DEFAULT_SHEET = "Results"
OUT_PATH = ROOT / "instrumentos" / "violin.py"

DYNAMICS = (
    "pppp",
    "ppp",
    "pp",
    "p",
    "mp",
    "mf",
    "f",
    "ff",
    "fff",
    "ffff",
)
NOTE_ORDER = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]


def _note_sort_key(note: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-G]#?)(-?\d+)", str(note).strip())
    if not m:
        return (99, 0)
    pc, octv = m.group(1), int(m.group(2))
    return (octv, NOTE_ORDER.index(pc) if pc in NOTE_ORDER else 99)


def _fmt(value: float) -> str:
    text = f"{float(value):.9g}"
    return text


def load_results_table(path: Path, sheet: str) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet!r} not in {path}; have {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise SystemExit(f"Empty sheet {sheet!r}")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    for required in ("note", *DYNAMICS):
        if required not in col:
            raise SystemExit(f"Missing column {required!r} on sheet {sheet!r}")

    table: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if not row or row[col["note"]] is None:
            continue
        status = row[col["status"]] if "status" in col else "ok"
        if status is not None and str(status).strip().lower() not in ("ok", ""):
            continue
        note = normalize_note_string(str(row[col["note"]]).strip())
        dyn_row: dict[str, float] = {}
        for dyn in DYNAMICS:
            val = row[col[dyn]]
            if val is None:
                raise SystemExit(f"Missing {dyn} for note {note}")
            dyn_row[dyn] = float(val)
        table[note] = dyn_row

    if not table:
        raise SystemExit("No usable rows loaded")
    return dict(sorted(table.items(), key=lambda kv: _note_sort_key(kv[0])))


def render_module(table: dict[str, dict[str, float]], *, source_path: Path, sheet: str) -> str:
    midis = []
    from microtonal import note_to_midi_strict

    for note in table:
        midis.append(int(round(note_to_midi_strict(note))))
    lo, hi = min(midis), max(midis)
    dyn_tuple = ", ".join(repr(d) for d in DYNAMICS)

    lines = [
        "# instrumentos/violin.py",
        '"""',
        "Violin instrument density module.",
        "",
        "The ``spectral_data`` table stores a committed 10-dynamic Combined Density",
        "Metric (CDM) ladder from the Dynamics_predicter ``Results`` sheet",
        "(IOWA + ORCH sustain anchors at pp/mf/ff; remaining levels committed from",
        "that workbook so runtime analysis does **not** re-run GPR / tail",
        "extrapolation for violin arco).",
        "",
        "Runtime analysis does not ingest audio; it maps notated pitch + dynamic to",
        "these pre-loaded acoustic metadata tables.",
        '"""',
        "",
        "from instrumentos.provenance import InstrumentSource",
        "",
        "INSTRUMENT_SOURCE = InstrumentSource(",
        '    source_type="external_acoustic_metadata",',
        "    citation=(",
        '        "Violin arco CDM ladder: IOWA+ORCH measured pp/mf/ff anchors with "',
        '        "committed Dynamics_predicter Results sheet values for all 10 "',
        '        "dynamic levels (not re-extrapolated at runtime)."',
        "    ),",
        "    source_url_or_identifier='docs/instrument_acoustic_sources.md#violin',",
        "    extraction_method=(",
        f'        "Committed full dynamic ladder from {source_path.name} / sheet {sheet!r}; "',
        '        "pitch lookup via MIDI-space spectral_lookup"',
        "    ),",
        f"    dynamic_levels=({dyn_tuple}),",
        f"    pitch_range=({lo}, {hi}),",
        '    uncertainty="medium",',
        '    version="2026-08-03",',
        '    source_technique="arco_sustain",',
        '    table_supported_techniques=("arco_sustain",),',
        ")",
        "",
        "import logging",
        "",
        "from utils.notes import normalize_note_string",
        "",
        'logger = logging.getLogger("violin")',
        "",
        "# Full 10-dynamic CDM ladder (Results sheet). Anchors pp/mf/ff match measured",
        "# IOWA+ORCH midpoints; other levels are workbook-committed (not runtime GPR).",
        "spectral_data = {",
    ]
    for note, dyns in table.items():
        parts = [f"{d!r}: {_fmt(dyns[d])}" for d in DYNAMICS]
        lines.append(f"    {note!r}: {{{', '.join(parts)}}},")
    lines.extend(
        [
            "}",
            "",
            "",
            "def calcular_densidade(nota, dinamica):",
            '    """Compute density from spectral CDM table (MIDI-space lookup, octave-safe)."""',
            "    from instrumentos.spectral_lookup import lookup_spectral_density",
            "",
            "    return lookup_spectral_density(",
            "        spectral_data,",
            "        nota,",
            "        dinamica,",
            "        logger=logger,",
            "        preprocess=normalize_note_string,",
            "    )",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--out", type=Path, default=OUT_PATH)
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise SystemExit(f"Workbook not found: {args.xlsx}")
    table = load_results_table(args.xlsx, args.sheet)
    text = render_module(table, source_path=args.xlsx, sheet=args.sheet)
    args.out.write_text(text, encoding="utf-8")
    print(f"Wrote {args.out} ({len(table)} pitches × {len(DYNAMICS)} dynamics from {args.sheet})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
